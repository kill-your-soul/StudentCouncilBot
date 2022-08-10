import os
from openpyxl import load_workbook
from async_sender import Mail
from ssl import create_default_context


context = create_default_context()
mail = Mail(
        hostname="smtp.gmail.com",
        # username="Bot ",
        port=465,
        use_tls=True,
        username="spbss.info@gmail.com",
        password=os.environ["GOOGLE_APP_PASSWORD"],
        tls_context=context,
    )
universities = []
unis = []
path = os.path.join("static", "emails.xlsx")
wb = load_workbook(path, read_only=True)
sheet = wb.active
for row in sheet.iter_rows(min_col=1, max_col=3):
    universities.append(
        {
            "university": row[0].value.lower(),
            "email": row[2].value.lower(),
            "name": row[1].value,
        }
    )
    unis.append(row[0].value.lower())


